# tests/reports/test_excel_builder.py
import os, tempfile
from backend.core.parser import load_excel
from backend.core.metrics import compute_all
from backend.reports.excel_builder import build_excel


def test_build_excel():
    b = compute_all(load_excel("resource/本周.xlsx"))
    p = os.path.join(tempfile.mkdtemp(), "r.xlsx")
    out = build_excel(b, [], {"week_compare": "测试", "daily_summary": "",
                              "procurement": "", "next_plan": ""}, p)
    assert os.path.exists(out)


def test_build_excel_has_all_sheets_and_native_chart():
    import openpyxl
    b = compute_all(load_excel("resource/本周.xlsx"))
    recent = [
        {"week_id": "W01", "销售额": 100.0, "销售毛利": 50.0,
         "销售毛利率": 0.5, "销售数量": 10},
        {"week_id": "W02", "销售额": 200.0, "销售毛利": 80.0,
         "销售毛利率": 0.4, "销售数量": 20},
    ]
    p = os.path.join(tempfile.mkdtemp(), "r.xlsx")
    build_excel(b, recent, {"week_compare": "测试", "daily_summary": "",
                            "procurement": "", "next_plan": ""}, p)
    wb = openpyxl.load_workbook(p)
    expected = {"汇总", "每日", "品牌", "平台", "店铺", "产品", "工厂", "AI文案"}
    assert expected.issubset(set(wb.sheetnames))
    # 汇总 sheet contains at least one native chart object
    assert len(wb["汇总"]._charts) >= 1
    # 店铺 sheet exists between 平台 and 产品
    names = wb.sheetnames
    assert names.index("店铺") > names.index("平台")
    assert names.index("店铺") < names.index("产品")
